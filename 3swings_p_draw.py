import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import openpyxl


wb=openpyxl.load_workbook('predict/swings_p.xlsx')
sht=wb['Sheet1']
line=[85,170,240,100]
series=["2023-wimbledon-2701","2023-usopen-1701","2021-frenchopen-1701","2021-ausopen-WS701"]
datas=[2,6,10,14]

font_path = 'Palatino Linotype/palatinolinotype_roman.ttf'
my_font = FontProperties(fname=font_path)

fig=plt.figure()

for n,i in enumerate(datas):
    mmtm_list=[]
    mmtm_p_list=[]

    for j in range(2,sht.max_row+1):
        if sht.cell(j,i).value==None:
            break
        mmtm_list.append(sht.cell(j,i).value)
        mmtm_p_list.append(sht.cell(j,i+1).value)

    swing_list_x=[]
    swing_list_y=[]
    swing_list_p_x=[]
    swing_list_p_y=[]
    for x in range(2,sht.max_row+1):
        if sht.cell(x,i+2).value==None:
            break
        if sht.cell(x,i+2).value==1:
            swing_list_x.append(x)
            swing_list_y.append(mmtm_list[x])
        if sht.cell(x,i+3).value==1:
            swing_list_p_x.append(x)
            swing_list_p_y.append(mmtm_p_list[x])

    ax=plt.subplot(2, 2, n+1)
    plt.scatter(swing_list_x,swing_list_y,s=9,zorder=3)
    plt.scatter(swing_list_p_x,swing_list_p_y,s=9,c='red',zorder=3)
    
    plt.plot(mmtm_list,linewidth=1, label='real')
    plt.plot(mmtm_p_list,linewidth=1,c='orange',label='predict')

    tmp=[0 if x!=None else None for x in mmtm_list]
    plt.plot(tmp, linewidth=1, color='grey')

    ax.axvline(x=line[n], color='r', linewidth=1, linestyle='--')

    ax.title.set_text(f'{series[n]}') 
    ax.title.set_fontproperties(my_font)  
    for label in (ax.get_xticklabels() + ax.get_yticklabels()):
        label.set_fontproperties(my_font)     
    #ax.legend(prop=my_font)

plt.show()
