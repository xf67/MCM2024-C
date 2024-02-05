import pandas as pd
from openpyxl import Workbook

data_all=pd.read_csv("data_w_mmtm.csv")
data=data_all[["DATA_NUM","set_no","game_no","point_no","new_mmtm1","new_mmtm2"]]    

out=[]
out_all=[]
workbook = Workbook()
sheet = workbook.active

col=1
row=1

for i in data["DATA_NUM"]:
    if i==0 or i!=7283 and data["point_no"][i+1]<data["point_no"][i]:
        col+=2
        row=1
    sheet.cell(row,col).value=data["new_mmtm1"][i]
    sheet.cell(row,col+1).value=data["new_mmtm2"][i]
    row+=1

workbook.save(filename="mmtm31.xlsx")



