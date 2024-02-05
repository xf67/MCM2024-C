import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import openpyxl


def main(data_out=False,draw=False,draw_index=False,index=1,draw_swings=False,read_swings=False):

    wb=openpyxl.load_workbook('new_mmtm_super_smooth.xlsx')
    sht=wb['Sheet1']
    sht2=wb['Sheet2']

    font_path = 'Palatino Linotype/palatinolinotype_roman.ttf'
    my_font = FontProperties(fname=font_path)

    if data_out:
        for i in range(1,32):
            #print(i)
            for j in range(1,sht.max_row+1):
                try:
                    sht2.cell(j,i).value=sht.cell(j,2*i-1).value-sht.cell(j,2*i).value
                except:
                    break
        wb.save('new_mmtm_super_smooth.xlsx')
    
    if draw:
        fig=plt.figure()
        #fig.set_size_inches(8,8)
        for i in range(1,31):
            mmtm_list=[]
            for j in range(1,sht2.max_row+1):
                mmtm_list.append(sht2.cell(j,i).value)
            ax=plt.subplot(6, 6, i)
            plt.plot(mmtm_list,linewidth=1)
            tmp=[0 if x!=None else None for x in mmtm_list]
            plt.plot(tmp, linewidth=1)
            ax.title.set_text(f'Series {i}') 
            ax.title.set_fontproperties(my_font)  
            for label in (ax.get_xticklabels() + ax.get_yticklabels()):
                label.set_fontproperties(my_font)     
        plt.show()

    if draw_index:
        fig=plt.figure()
        i=index
        if i>31:
            print("i must <= 31")
            exit()
        mmtm_list=[]
        for j in range(1,sht2.max_row+1):
            mmtm_list.append(sht2.cell(j,i).value)
        ax=plt.subplot(1, 1, 1)
        plt.plot(mmtm_list,linewidth=1)
        tmp=[0 if x!=None else None for x in mmtm_list]
        plt.plot(tmp, linewidth=1)
        ax.title.set_text(f'Series {i}') 
        ax.title.set_fontproperties(my_font)  
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontproperties(my_font)
        plt.show()
    
    if draw_swings:
        wbb=openpyxl.load_workbook('swings.xlsx')
        sss=wbb['Sheet1']
        fig=plt.figure()
        #fig.set_size_inches(8,8)
        for i in range(1,31):
            mmtm_list=[]
            for j in range(1,sht2.max_row+1):
                mmtm_list.append(sht2.cell(j,i).value)
            
            ax=plt.subplot(5, 6, i)

            swing_list_x=[]
            swing_list_y=[]
            for x in range(2,sss.max_row):
                if sss.cell(x,i).value==None:
                    break
                swing_list_x.append(sss.cell(x,i).value)
                swing_list_y.append(mmtm_list[sss.cell(x,i).value])

            plt.scatter(swing_list_x,swing_list_y,s=10,c='red',zorder=3)
            
            plt.plot(mmtm_list,linewidth=1)

            tmp=[0 if x!=None else None for x in mmtm_list]
            plt.plot(tmp, linewidth=1, color='grey')
            
            
            ax.title.set_text(f'Series {i}') 
            ax.title.set_fontproperties(my_font)  
            for label in (ax.get_xticklabels() + ax.get_yticklabels()):
                label.set_fontproperties(my_font)     
        plt.show()


if __name__ == '__main__':
    main(data_out=False,draw=False,draw_index=False,index=9,draw_swings=True,read_swings=False)